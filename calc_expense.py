import openpyxl as xl
from openpyxl.chart import Reference, BarChart


exp_total_dict = dict()


def calculate_total(filename):
    wb = xl.load_workbook(filename)
    sheet_names = wb.sheetnames
    ws = wb.create_sheet('Summary')
    rows = [("Month", "Total Expenditure")]
    for row in rows:
        ws.append(row)
    for sheet_name in sheet_names:
        total_exp = 0
        sheet = wb[sheet_name]
        for row in range(1, sheet.max_row+1):
            if str(sheet.cell(row, 1).value).upper() == "GYM":
                total_exp += sheet.cell(row, 3).value
        rows = [(sheet_name, total_exp)]
        for row in rows:
            ws.append(row)
        wb.save('SampleData.xlsx')
        exp_total_dict[sheet_name] = total_exp
    return exp_total_dict


