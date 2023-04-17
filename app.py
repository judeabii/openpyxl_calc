from pathlib import Path
import openpyxl as xl
from openpyxl.chart import Reference, BarChart
from calc_expense import calculate_total

gym_total_dict = {}
path = Path()
for file in path.glob('SampleData.xlsx'):
    gym_total_dict = calculate_total(file)

print(gym_total_dict)

wb = xl.load_workbook('SampleData.xlsx')
ws = wb['Summary']

data = Reference(ws, min_row=1, max_row=ws.max_row, min_col=2, max_col=2)
titles = Reference(ws, min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
chart = BarChart()
chart.title = "Gym Expense Chart"
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)

ws.add_chart(chart, "H2")
wb.save("SampleData.xlsx")
